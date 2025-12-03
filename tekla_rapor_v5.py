# -*- coding: utf-8 -*-
import pandas as pd
import re
import os
from tkinter import Tk, filedialog, messagebox, simpledialog, scrolledtext, ttk
import tkinter as tk
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')
from tkinter import StringVar, IntVar, BooleanVar

# Openpyxl'i import et
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_KURULU = True
except ImportError:
    OPENPYXL_KURULU = False
    print("UYARI: openpyxl kurulu değil, basit kaydetme kullanılacak")

class TeklaRaporIsleyici:
    def __init__(self):
        self.veri_df = None
        self.ozet_df = None
        self.dosya_yolu = ""
        self.ham_veriler = None
        self.dosya_adi = ""
        self.kesim_verileri = None
        self.kesim_optimizasyon_sonucu = None
        self.uretilen_excel_yolu = None  # Üretilen excel yolunu sakla
        
    def dosya_sec(self):
        """Dosya seçme dialog'u"""
        Tk().withdraw()
        dosya_turleri = [
            ("Excel files", "*.xls *.xlsx *.xlsm"),
            ("Text files", "*.txt"),
            ("XSR files", "*.xsr"),
            ("CSV files", "*.csv"),
            ("All files", "*.*")
        ]
        
        self.dosya_yolu = filedialog.askopenfilename(
            title="Tekla Rapor Dosyasını Seçin",
            filetypes=dosya_turleri
        )
        
        if self.dosya_yolu:
            self.dosya_adi = os.path.splitext(os.path.basename(self.dosya_yolu))[0]
            # Sekme adı için temizleme (max 31 karakter, özel karakterler yok)
            self.dosya_adi = re.sub(r'[\\/*?:\[\]]', '', self.dosya_adi)[:31]
        
        return self.dosya_yolu
    
    def plaka_kalinligini_bul(self, plaka_adi):
        """PL ifadelerinden kalınlık değerini çıkar"""
        if pd.isna(plaka_adi):
            return ""
        
        plaka_str = str(plaka_adi).upper().strip()
        
        if not plaka_str.startswith('PL'):
            return ""
        
        plaka_bilgi = plaka_str[2:]
        
        ayraclar = ['*', 'x', 'X']
        ayrac_bulundu = False
        secilen_ayrac = ''
        ayrac_index = -1
        
        for ayrac in ayraclar:
            if ayrac in plaka_bilgi:
                ayrac_bulundu = True
                secilen_ayrac = ayrac
                ayrac_index = plaka_bilgi.index(ayrac)
                break
        
        if ayrac_bulundu:
            sol_taraf = plaka_bilgi[:ayrac_index]
            sag_taraf = plaka_bilgi[ayrac_index + 1:]
            
            sol_sayi = self.sayisal_degeri_al(sol_taraf)
            sag_sayi = self.sayisal_degeri_al(sag_taraf)
            
            if sol_sayi > 0 and sag_sayi > 0:
                kalinlik = min(sol_sayi, sag_sayi)
            elif sol_sayi > 0:
                kalinlik = sol_sayi
            elif sag_sayi > 0:
                kalinlik = sag_sayi
            else:
                kalinlik = 0
        else:
            kalinlik = self.sayisal_degeri_al(plaka_bilgi)
        
        if kalinlik > 0:
            return f"{kalinlik} mm Sac"
        else:
            return ""
    
    def sayisal_degeri_al(self, metin):
        """Metinden sayısal değer çıkar"""
        if pd.isna(metin):
            return 0.0
        
        metin_str = str(metin)
        sayi_metni = ""
        ondalik_ayrac = False
        
        for karakter in metin_str:
            if karakter.isdigit():
                sayi_metni += karakter
            elif karakter in ['.', ','] and not ondalik_ayrac:
                sayi_metni += '.'
                ondalik_ayrac = True
        
        try:
            return float(sayi_metni) if sayi_metni else 0.0
        except:
            return 0.0
    
    def kesim_verilerini_otomatik_bul(self, excel_yolu=None):
        """Üretilen Excel'den otomatik kesim verilerini bul"""
        try:
            if excel_yolu is None:
                if self.uretilen_excel_yolu:
                    excel_yolu = self.uretilen_excel_yolu
                else:
                    # Son üretilen excel'i bulmaya çalış
                    klasor = os.path.dirname(self.dosya_yolu) if self.dosya_yolu else "."
                    excel_dosyalari = [f for f in os.listdir(klasor) if f.startswith("Tekla_Rapor_") and f.endswith(".xlsx")]
                    
                    if excel_dosyalari:
                        # En yeni dosyayı al
                        excel_dosyalari.sort(key=lambda x: os.path.getmtime(os.path.join(klasor, x)), reverse=True)
                        excel_yolu = os.path.join(klasor, excel_dosyalari[0])
                    else:
                        messagebox.showwarning("Uyarı", "Üretilmiş Excel dosyası bulunamadı!")
                        return False
            
            # Excel'i oku
            excel_data = pd.read_excel(excel_yolu, sheet_name='Data')
            
            # Kesim verilerini ayıkla (Length(mm) sütunundaki veriler)
            kesim_verileri = []
            
            # Length(mm) sütunundaki benzersiz değerleri ve adetlerini bul
            if 'Length(mm)' in excel_data.columns:
                length_values = excel_data['Length(mm)'].dropna()
                
                # Benzersiz değerleri ve adetlerini say
                for value in length_values.unique():
                    if pd.notna(value):
                        try:
                            boy = float(value)
                            if boy > 0:
                                adet = int((length_values == value).sum())
                                kesim_verileri.append({
                                    'Adet': adet,
                                    'Boy': boy,
                                    'Toplam Uzunluk': adet * boy
                                })
                        except:
                            continue
            
            if kesim_verileri:
                self.kesim_verileri = pd.DataFrame(kesim_verileri)
                
                # Verileri kaydet
                self.uretilen_excel_yolu = excel_yolu
                
                return True
            else:
                # Alternatif: Qty ve Length(mm) sütunlarını kullan
                if 'Qty' in excel_data.columns and 'Length(mm)' in excel_data.columns:
                    kesim_verileri = []
                    for idx, row in excel_data.iterrows():
                        try:
                            if pd.notna(row['Qty']) and pd.notna(row['Length(mm)']):
                                adet = int(float(row['Qty']))
                                boy = float(row['Length(mm)'])
                                if adet > 0 and boy > 0:
                                    kesim_verileri.append({
                                        'Adet': adet,
                                        'Boy': boy,
                                        'Toplam Uzunluk': adet * boy
                                    })
                        except:
                            continue
                    
                    if kesim_verileri:
                        self.kesim_verileri = pd.DataFrame(kesim_verileri)
                        self.uretilen_excel_yolu = excel_yolu
                        return True
            
            messagebox.showwarning("Uyarı", "Data sayfasında kesim verileri bulunamadı!")
            return False
            
        except Exception as e:
            messagebox.showerror("Hata", f"Kesim verileri okuma sırasında hata: {str(e)}")
            return False
    
    def benzersiz_yap(self):
        """Kesim planı için benzersizleştirme işlemi"""
        if self.kesim_verileri is None or self.kesim_verileri.empty:
            # Önce otomatik olarak kesim verilerini bulmaya çalış
            if not self.kesim_verilerini_otomatik_bul():
                messagebox.showwarning("Uyarı", "Kesim verileri yüklenmemiş!")
                return False
        
        try:
            # C sütunundaki boy değerlerine göre benzersizleştirme
            benzersiz_df = pd.DataFrame(columns=['Adet', 'Boy', 'Toplam Uzunluk'])
            
            # Boy değerlerine göre grupla ve adetleri topla
            for boy in self.kesim_verileri['Boy'].unique():
                toplam_adet = self.kesim_verileri[self.kesim_verileri['Boy'] == boy]['Adet'].sum()
                toplam_uzunluk = toplam_adet * boy
                benzersiz_df = pd.concat([benzersiz_df, pd.DataFrame([{
                    'Adet': int(toplam_adet),
                    'Boy': float(boy),
                    'Toplam Uzunluk': float(toplam_uzunluk)
                }])], ignore_index=True)
            
            # Büyükten küçüğe sırala
            benzersiz_df = benzersiz_df.sort_values('Boy', ascending=False).reset_index(drop=True)
            
            # Orijinal veriyi güncelle
            self.kesim_verileri = benzersiz_df
            
            # Benzersiz verileri yeni bir Excel sayfasına kaydet
            self.benzersiz_verileri_excele_kaydet()
            
            return True
            
        except Exception as e:
            messagebox.showerror("Hata", f"Benzersizleştirme sırasında hata: {str(e)}")
            return False
    
    def benzersiz_verileri_excele_kaydet(self):
        """Benzersiz verileri yeni bir Excel sayfasına kaydet"""
        try:
            if self.uretilen_excel_yolu and os.path.exists(self.uretilen_excel_yolu):
                # Mevcut Excel dosyasını aç
                wb = openpyxl.load_workbook(self.uretilen_excel_yolu)
                
                # Benzersiz veriler sayfasını oluştur (varsa sil)
                if 'Benzersiz_Kesim' in wb.sheetnames:
                    ws = wb['Benzersiz_Kesim']
                    wb.remove(ws)
                
                ws = wb.create_sheet(title='Benzersiz_Kesim')
                
                # Başlıklar
                ws.cell(row=1, column=1, value="BENZERSİZ KESİM VERİLERİ")
                ws.cell(row=1, column=1).font = Font(size=14, bold=True)
                ws.merge_cells('A1:C1')
                
                ws.cell(row=2, column=1, value=f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
                
                # Alt başlıklar
                ws.cell(row=4, column=1, value="Adet")
                ws.cell(row=4, column=2, value="Boy (mm)")
                ws.cell(row=4, column=3, value="Toplam Uzunluk (mm)")
                
                for col in range(1, 4):
                    ws.cell(row=4, column=col).font = Font(bold=True)
                    ws.cell(row=4, column=col).fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                
                # Veriler
                start_row = 5
                for idx, row in self.kesim_verileri.iterrows():
                    ws.cell(row=start_row + idx, column=1, value=int(row['Adet']))
                    ws.cell(row=start_row + idx, column=2, value=float(row['Boy']))
                    ws.cell(row=start_row + idx, column=3, value=float(row['Toplam Uzunluk']))
                
                # Toplam satırı
                total_row = start_row + len(self.kesim_verileri) + 1
                ws.cell(row=total_row, column=1, value="TOPLAM")
                ws.cell(row=total_row, column=2, value="")
                ws.cell(row=total_row, column=3, value=f"=SUM(C{start_row}:C{total_row-1})")
                
                for col in [1, 3]:
                    ws.cell(row=total_row, column=col).font = Font(bold=True)
                    ws.cell(row=total_row, column=col).fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                
                # Sütun genişlikleri
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Sayfayı kaydet
                wb.save(self.uretilen_excel_yolu)
                
                return True
            
        except Exception as e:
            print(f"Benzersiz verileri Excel'e kaydetme hatası: {str(e)}")
            return False
    
    def kesim_optimizasyonu_hesapla(self, stok_boy=12000, min_kullanilabilir_fire=400):
        """Kesim optimizasyonu hesapla"""
        if self.kesim_verileri is None or self.kesim_verileri.empty:
            # Önce otomatik olarak kesim verilerini bulmaya çalış
            if not self.kesim_verilerini_otomatik_bul():
                messagebox.showwarning("Uyarı", "Kesim verileri yüklenmemiş!")
                return None
        
        try:
            # Kesim verilerini hazırla
            parcalar = []
            for idx, row in self.kesim_verileri.iterrows():
                for _ in range(int(row['Adet'])):
                    parcalar.append({
                        'Boy': float(row['Boy']),
                        'Kullanildi': False,
                        'KesimNo': 0
                    })
            
            # Parçaları büyükten küçüğe sırala
            parcalar.sort(key=lambda x: x['Boy'], reverse=True)
            
            # Kesim planı oluştur
            kesim_planlari = []
            kesim_no = 1
            
            # Kullanılmayan parçaları takip et
            kullanilmayan_parcalar = parcalar.copy()
            
            while True:
                # Henüz kullanılmayan parçaları bul
                kalan_parcalar = [p for p in parcalar if not p['Kullanildi']]
                
                if not kalan_parcalar:
                    break
                
                # Yeni bir stok başlat
                stok_uzunluk = stok_boy
                kesim_listesi = []
                
                # En büyük parçadan başlayarak yerleştir
                for parca in kalan_parcalar:
                    if not parca['Kullanildi'] and parca['Boy'] <= stok_uzunluk:
                        kesim_listesi.append(parca['Boy'])
                        stok_uzunluk -= parca['Boy']
                        parca['Kullanildi'] = True
                        parca['KesimNo'] = kesim_no
                
                # Fire hesapla
                fire = stok_uzunluk
                
                # Kesim planına ekle
                kesim_planlari.append({
                    'Kesim No': kesim_no,
                    'Stok Boy': stok_boy,
                    'Kesimler': kesim_listesi,
                    'Toplam Kesim': sum(kesim_listesi),
                    'Fire': fire,
                    'Kullanım %': (sum(kesim_listesi) / stok_boy) * 100 if stok_boy > 0 else 0
                })
                
                kesim_no += 1
                
                # Güvenlik için maksimum kesim sayısı
                if kesim_no > 100:
                    break
            
            # Optimizasyon sonuçlarını dataframe'e çevir
            optimizasyon_df = pd.DataFrame(kesim_planlari)
            
            # Özet bilgileri hesapla
            toplam_fire = optimizasyon_df['Fire'].sum()
            toplam_kullanilan = optimizasyon_df['Toplam Kesim'].sum()
            ortalama_kullanim = optimizasyon_df['Kullanım %'].mean()
            kullanilan_stok_sayisi = len(optimizasyon_df)
            
            # Kullanılmayan parçaları bul
            kullanilmayan = [p for p in parcalar if not p['Kullanildi']]
            
            # Sonuçları kaydet
            self.kesim_optimizasyon_sonucu = {
                'optimizasyon_df': optimizasyon_df,
                'toplam_fire': toplam_fire,
                'toplam_kullanilan': toplam_kullanilan,
                'ortalama_kullanim': ortalama_kullanim,
                'kullanilan_stok_sayisi': kullanilan_stok_sayisi,
                'kullanilmayan_parcalar': kullanilmayan,
                'stok_boy': stok_boy,
                'min_fire': min_kullanilabilir_fire
            }
            
            # Optimizasyon sonuçlarını Excel'e kaydet
            self.optimizasyon_sonuclarini_excele_kaydet()
            
            return self.kesim_optimizasyon_sonucu
            
        except Exception as e:
            messagebox.showerror("Hata", f"Optimizasyon hesaplama sırasında hata: {str(e)}")
            return None
    
    def optimizasyon_sonuclarini_excele_kaydet(self):
          def optimizasyon_sonuclarini_excele_kaydet(self):
        """Optimizasyon sonuçlarını Excel'e kaydet - Malzeme tipine göre gruplu"""
        try:
            if self.uretilen_excel_yolu and os.path.exists(self.uretilen_excel_yolu) and self.kesim_optimizasyon_sonucu:
                # Mevcut Excel dosyasını aç
                wb = openpyxl.load_workbook(self.uretilen_excel_yolu)
                
                # DATA sayfasındaki malzeme bilgilerini oku
                if 'Data' in wb.sheetnames:
                    data_df = pd.read_excel(self.uretilen_excel_yolu, sheet_name='Data')
                else:
                    data_df = pd.DataFrame()
                
                # Optimizasyon sayfasını oluştur (varsa sil)
                if 'Kesim_Optimizasyon' in wb.sheetnames:
                    ws = wb['Kesim_Optimizasyon']
                    wb.remove(ws)
                
                ws = wb.create_sheet(title='Kesim_Optimizasyon')
                
                sonuc = self.kesim_optimizasyon_sonucu
                
                # ========== 1. ÖZET BİLGİLER ==========
                ws.cell(row=1, column=1, value="KESİM OPTİMİZASYON SONUÇLARI")
                ws.cell(row=1, column=1).font = Font(size=16, bold=True)
                ws.merge_cells('A1:F1')
                
                # Tarih
                ws.cell(row=2, column=1, value=f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
                
                # Özet bilgiler
                ws.cell(row=4, column=1, value="ÖZET BİLGİLER")
                ws.cell(row=4, column=1).font = Font(bold=True)
                
                ozet_bilgiler = [
                    ("Stok Boyu:", f"{sonuc['stok_boy']} mm"),
                    ("Minimum Fire:", f"{sonuc['min_fire']} mm"),
                    ("Kullanılan Stok:", f"{sonuc['kullanilan_stok_sayisi']} adet"),
                    ("Toplam Kullanılan:", f"{sonuc['toplam_kullanilan']:,.0f} mm"),
                    ("Toplam Fire:", f"{sonuc['toplam_fire']:,.0f} mm"),
                    ("Ortalama Kullanım:", f"{sonuc['ortalama_kullanim']:.1f}%"),
                ]
                
                for i, (label, value) in enumerate(ozet_bilgiler, start=5):
                    ws.cell(row=i, column=1, value=label)
                    ws.cell(row=i, column=2, value=value)
                
                # ========== 2. MALZEME TİPİNE GÖRE GRUPLAMA ==========
                if not data_df.empty and 'Size' in data_df.columns:
                    start_row = len(ozet_bilgiler) + 8
                    
                    ws.cell(row=start_row, column=1, value="MALZEME TİPİNE GÖRE KESİM DAĞILIMI")
                    ws.cell(row=start_row, column=1).font = Font(size=14, bold=True)
                    ws.merge_cells(f'A{start_row}:F{start_row}')
                    
                    # Başlıklar
                    headers = ["Malzeme Tipi", "Adet", "Toplam Boy (mm)", "Stok Gereksinimi", "Fire (mm)", "Kullanım %"]
                    header_row = start_row + 2
                    for col, header in enumerate(headers, start=1):
                        ws.cell(row=header_row, column=col, value=header)
                        ws.cell(row=header_row, column=col).font = Font(bold=True)
                        ws.cell(row=header_row, column=col).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    
                    # Malzeme tiplerini grupla
                    malzeme_gruplari = {}
                    for idx, row in data_df.iterrows():
                        size = row['Size']
                        if pd.notna(row.get('Length(mm)')):
                            length = row['Length(mm)']
                            qty = row['Qty'] if pd.notna(row.get('Qty')) else 1
                            
                            if size not in malzeme_gruplari:
                                malzeme_gruplari[size] = {
                                    'adet': 0,
                                    'toplam_boy': 0,
                                    'boylar': []
                                }
                            
                            malzeme_gruplari[size]['adet'] += qty
                            malzeme_gruplari[size]['toplam_boy'] += length * qty
                            for _ in range(int(qty)):
                                malzeme_gruplari[size]['boylar'].append(length)
                    
                    # Her malzeme grubu için optimizasyon hesapla
                    data_row = header_row + 1
                    for malzeme_tipi, veri in malzeme_gruplari.items():
                        # Bu malzeme tipi için kesim optimizasyonu
                        boylar = veri['boylar']
                        boylar.sort(reverse=True)
                        
                        kesim_planlari = []
                        stok_boy = sonuc['stok_boy']
                        
                        while boylar:
                            kalan_stok = stok_boy
                            kesilenler = []
                            
                            for boy in boylar[:]:  # Copy of list
                                if boy <= kalan_stok:
                                    kesilenler.append(boy)
                                    kalan_stok -= boy
                                    boylar.remove(boy)
                            
                            if kesilenler:
                                fire = kalan_stok
                                kullanim_orani = ((stok_boy - fire) / stok_boy) * 100 if stok_boy > 0 else 0
                                
                                kesim_planlari.append({
                                    'kesilenler': kesilenler,
                                    'fire': fire,
                                    'kullanim': kullanim_orani
                                })
                            else:
                                break
                        
                        # Hesaplamalar
                        toplam_stok = len(kesim_planlari)
                        toplam_fire = sum(k['fire'] for k in kesim_planlari)
                        ortalama_kullanim = sum(k['kullanim'] for k in kesim_planlari) / len(kesim_planlari) if kesim_planlari else 0
                        
                        # Excel'e yaz
                        ws.cell(row=data_row, column=1, value=malzeme_tipi)
                        ws.cell(row=data_row, column=2, value=veri['adet'])
                        ws.cell(row=data_row, column=3, value=veri['toplam_boy'])
                        ws.cell(row=data_row, column=4, value=toplam_stok)
                        ws.cell(row=data_row, column=5, value=toplam_fire)
                        ws.cell(row=data_row, column=6, value=f"{ortalama_kullanim:.1f}%")
                        
                        data_row += 1
                    
                    # Toplam satırı
                    ws.cell(row=data_row, column=1, value="TOPLAM")
                    ws.cell(row=data_row, column=2, value=f"=SUM(B{header_row+1}:B{data_row-1})")
                    ws.cell(row=data_row, column=3, value=f"=SUM(C{header_row+1}:C{data_row-1})")
                    ws.cell(row=data_row, column=4, value=f"=SUM(D{header_row+1}:D{data_row-1})")
                    ws.cell(row=data_row, column=5, value=f"=SUM(E{header_row+1}:E{data_row-1})")
                    ws.cell(row=data_row, column=6, value="")
                    
                    for col in range(1, 7):
                        ws.cell(row=data_row, column=col).font = Font(bold=True)
                        ws.cell(row=data_row, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # ========== 3. DETAYLI KESİM PLANI ==========
                detay_start_row = data_row + 3
                ws.cell(row=detay_start_row, column=1, value="DETAYLI KESİM PLANI")
                ws.cell(row=detay_start_row, column=1).font = Font(size=14, bold=True)
                ws.merge_cells(f'A{detay_start_row}:F{detay_start_row}')
                
                # Başlıklar
                detay_headers = ["Kesim No", "Stok Boy (mm)", "Kesimler", "Toplam Kesim (mm)", "Fire (mm)", "Kullanım %"]
                detay_header_row = detay_start_row + 2
                for col, header in enumerate(detay_headers, start=1):
                    ws.cell(row=detay_header_row, column=col, value=header)
                    ws.cell(row=detay_header_row, column=col).font = Font(bold=True)
                    ws.cell(row=detay_header_row, column=col).fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                
                # Veriler
                for idx, kesim in sonuc['optimizasyon_df'].iterrows():
                    row = detay_header_row + idx + 1
                    ws.cell(row=row, column=1, value=kesim['Kesim No'])
                    ws.cell(row=row, column=2, value=kesim['Stok Boy'])
                    
                    # Kesimleri virgülle ayırarak yaz
                    kesim_str = ", ".join([str(int(k)) for k in kesim['Kesimler']])
                    ws.cell(row=row, column=3, value=kesim_str)
                    
                    ws.cell(row=row, column=4, value=kesim['Toplam Kesim'])
                    ws.cell(row=row, column=5, value=kesim['Fire'])
                    ws.cell(row=row, column=6, value=f"{kesim['Kullanım %']:.1f}%")
                
                # Kullanılmayan parçalar
                if sonuc['kullanilmayan_parcalar']:
                    unused_row = detay_header_row + len(sonuc['optimizasyon_df']) + 3
                    ws.cell(row=unused_row, column=1, value="KULLANILMAYAN PARÇALAR:")
                    ws.cell(row=unused_row, column=1).font = Font(bold=True)
                    
                    for i, parca in enumerate(sonuc['kullanilmayan_parcalar'], start=1):
                        ws.cell(row=unused_row + i, column=1, value=f"- {parca['Boy']} mm")
                
                # ========== 4. FORMATLAMA ==========
                # Sütun genişlikleri
                column_widths = [35, 12, 15, 18, 12, 12]
                for i, width in enumerate(column_widths, start=1):
                    column_letter = get_column_letter(i)
                    ws.column_dimensions[column_letter].width = width
                
                # Hizalama
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
                    for cell in row:
                        if cell.column in [2, 3, 4, 5, 6]:  # Sayısal sütunlar
                            cell.alignment = Alignment(horizontal='right')
                        else:
                            cell.alignment = Alignment(horizontal='left')
                
                # Auto-filter
                ws.auto_filter.ref = f"A{detay_header_row}:F{detay_header_row + len(sonuc['optimizasyon_df'])}"
                
                # Kaydet
                wb.save(self.uretilen_excel_yolu)
                
                print(f"✅ Optimizasyon raporu güncellendi: Malzeme tipine göre gruplandırıldı")
                return True
            
        except Exception as e:
            print(f"❌ Optimizasyon sonuçlarını Excel'e kaydetme hatası: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
                
                return True
            
        except Exception as e:
            print(f"Optimizasyon sonuçlarını Excel'e kaydetme hatası: {str(e)}")
            return False
    
    def kesim_optimizasyon_raporu_olustur(self):
        """Kesim optimizasyonu raporu oluştur"""
        if self.kesim_optimizasyon_sonucu is None:
            messagebox.showwarning("Uyarı", "Önce optimizasyon hesaplanmalı!")
            return None
        
        rapor = []
        sonuc = self.kesim_optimizasyon_sonucu
        
        rapor.append("=" * 60)
        rapor.append("KESİM OPTİMİZASYON RAPORU")
        rapor.append("=" * 60)
        rapor.append(f"Stok Boyu: {sonuc['stok_boy']} mm")
        rapor.append(f"Minimum Kullanılabilir Fire: {sonuc['min_fire']} mm")
        rapor.append(f"Kullanılan Stok Sayısı: {sonuc['kullanilan_stok_sayisi']} adet")
        rapor.append(f"Toplam Kullanılan Uzunluk: {sonuc['toplam_kullanilan']:,.0f} mm")
        rapor.append(f"Toplam Fire: {sonuc['toplam_fire']:,.0f} mm")
        rapor.append(f"Ortalama Kullanım Oranı: {sonuc['ortalama_kullanim']:.1f}%")
        rapor.append("")
        
        # Kullanılmayan parçalar
        if sonuc['kullanilmayan_parcalar']:
            rapor.append("KULLANILMAYAN PARÇALAR:")
            for parca in sonuc['kullanilmayan_parcalar']:
                rapor.append(f"  - {parca['Boy']} mm")
        
        rapor.append("")
        rapor.append("KESİM PLANI DETAYLARI:")
        rapor.append("-" * 60)
        
        for idx, kesim in sonuc['optimizasyon_df'].iterrows():
            rapor.append(f"KESİM {kesim['Kesim No']}:")
            rapor.append(f"  Stok Boy: {kesim['Stok Boy']} mm")
            kesim_detay = ", ".join([f"{int(k)} mm" for k in kesim['Kesimler']])
            rapor.append(f"  Kesimler: {kesim_detay}")
            rapor.append(f"  Toplam Kesim: {kesim['Toplam Kesim']} mm")
            rapor.append(f"  Fire: {kesim['Fire']} mm")
            rapor.append(f"  Kullanım: {kesim['Kullanım %']:.1f}%")
            rapor.append("")
        
        return "\n".join(rapor)
    
    def kaynak_sayfa_olustur(self, workbook):
        """Kaynak dosyayı yeni bir sayfaya kopyala"""
        if self.ham_veriler is None or (hasattr(self.ham_veriler, 'empty') and self.ham_veriler.empty):
            return False
        
        try:
            sayfa_adi = self.dosya_adi if self.dosya_adi else "Kaynak"
            
            sheet_names = workbook.sheetnames
            original_name = sayfa_adi
            counter = 1
            
            while sayfa_adi in sheet_names:
                sayfa_adi = f"{original_name}_{counter}"
                counter += 1
                if counter > 10:
                    sayfa_adi = f"Kaynak_{datetime.now().strftime('%H%M%S')}"
                    break
            
            kaynak_sheet = workbook.create_sheet(title=sayfa_adi)
            
            if isinstance(self.ham_veriler, pd.DataFrame):
                if len(self.ham_veriler.columns) == 1 and self.ham_veriler.columns[0] == 'Raw':
                    for i, value in enumerate(self.ham_veriler['Raw'], start=1):
                        cell_value = str(value).strip() if pd.notna(value) else ""
                        kaynak_sheet.cell(row=i, column=1, value=cell_value)
                else:
                    for r_idx, row in enumerate(self.ham_veriler.values, start=1):
                        for c_idx, value in enumerate(row, start=1):
                            cell_value = value if pd.notna(value) else ""
                            kaynak_sheet.cell(row=r_idx, column=c_idx, value=cell_value)
            else:
                for i, value in enumerate(self.ham_veriler, start=1):
                    cell_value = str(value).strip() if pd.notna(value) else ""
                    kaynak_sheet.cell(row=i, column=1, value=cell_value)
            
            if OPENPYXL_KURULU:
                kaynak_sheet.insert_rows(1)
                kaynak_sheet.cell(row=1, column=1, value=f"KAYNAK DOSYA: {os.path.basename(self.dosya_yolu)}")
                kaynak_sheet.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
                kaynak_sheet.cell(row=1, column=1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                kaynak_sheet.cell(row=2, column=1, value=f"Kopyalama Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
                kaynak_sheet.cell(row=2, column=1).font = Font(italic=True, size=9)
                
                kaynak_sheet.insert_rows(4)
            
            workbook._sheets.insert(0, workbook._sheets.pop(workbook._sheets.index(kaynak_sheet)))
            
            return True
            
        except Exception as e:
            print(f"Kaynak sayfa oluşturma hatası: {str(e)}")
            return False
    
    def tum_islemleri_yap(self):
        """Tüm işlemleri tek seferde yap"""
        sonuclar = []
        
        # 1. Dosya seç
        self.dosya_yolu = self.dosya_sec()
        if not self.dosya_yolu:
            return ["İşlem iptal edildi: Dosya seçilmedi"]
        
        dosya_adi = os.path.basename(self.dosya_yolu)
        sonuclar.append(f"📂 Dosya seçildi: {dosya_adi}")
        
        # 2. Dosyayı aç ve oku
        try:
            dosya_uzanti = os.path.splitext(self.dosya_yolu)[1].lower()
            
            if dosya_uzanti in ['.xls', '.xlsx', '.xlsm']:
                self.ham_veriler = pd.read_excel(self.dosya_yolu, header=None)
                df = self.ham_veriler.copy()
                okunan_satir = len(df)
                dosya_tipi = "Excel"
            else:
                with open(self.dosya_yolu, 'r', encoding='utf-8', errors='ignore') as f:
                    lines = f.readlines()
                self.ham_veriler = pd.DataFrame(lines, columns=['Raw'])
                df = self.ham_veriler.copy()
                okunan_satir = len(df)
                dosya_tipi = "Text/XSR"
            
            sonuclar.append(f"📄 {dosya_tipi} dosyası açıldı: {okunan_satir} satır")
            
        except Exception as e:
            return [f"❌ HATA: Dosya açılamadı - {str(e)}"]
        
        # 3. Verileri işle (PL dönüşümleri ile)
        processed_data = []
        
        for idx, row in df.iterrows():
            try:
                if len(df.columns) > 1:
                    cell_value = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
                else:
                    cell_value = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
            except (IndexError, KeyError):
                continue
            
            cell_value = cell_value.strip()
            
            if not cell_value:
                continue
            
            cell_upper = cell_value.upper()
            skip_keywords = ['TEKLA', 'SIZE', '-----', 'TOTAL', 'RAKAMLAR', 'METRAJ', 'TÜM', 'ALL']
            if any(kelime in cell_upper for kelime in skip_keywords):
                continue
            
            if cell_value and not cell_value[0].isalpha():
                continue
            
            parts = cell_value.split()
            
            if len(parts) >= 6:
                try:
                    size_adi = parts[0]
                    
                    plaka_donusum = self.plaka_kalinligini_bul(size_adi)
                    if plaka_donusum:
                        size_adi = plaka_donusum
                    
                    grade = parts[1]
                    
                    qty = float(str(parts[2]).replace(',', '.'))
                    length = float(str(parts[3]).replace(',', '.'))
                    area = float(str(parts[4]).replace(',', '.'))
                    weight = float(str(parts[5]).replace(',', '.'))
                    
                    total_kg = qty * weight
                    total_mm = qty * length
                    
                    processed_data.append({
                        'Size': size_adi,
                        'Grade': grade,
                        'Qty': qty,
                        'Length(mm)': length,
                        'Area(m²)': area,
                        'Weight(kg)': weight,
                        'Total Kg': total_kg,
                        'Total mm': total_mm
                    })
                except (ValueError, IndexError, TypeError):
                    continue
        
        if not processed_data:
            return ["❌ HATA: İşlenebilir veri bulunamadı!"]
        
        self.veri_df = pd.DataFrame(processed_data)
        islenen_satir = len(self.veri_df)
        sonuclar.append(f"✅ Veri işlendi: {islenen_satir} satır")
        
        plaka_sayisi = self.veri_df['Size'].str.contains('mm Sac', na=False).sum()
        if plaka_sayisi > 0:
            sonuclar.append(f"🔨 {plaka_sayisi} adet PL dönüşümü yapıldı")
        
        # 4. Özet tablo oluştur
        ozet_data = []
        for size, group in self.veri_df.groupby('Size'):
            ozet_data.append({
                'Size': size,
                'Toplam mm': group['Total mm'].sum(),
                'Toplam Kg': group['Total Kg'].sum()
            })
        
        self.ozet_df = pd.DataFrame(ozet_data).sort_values('Size').reset_index(drop=True)
        farkli_malzeme = len(self.ozet_df)
        sonuclar.append(f"📊 Özet tablo: {farkli_malzeme} farklı malzeme")
        
        # 5. Excel'e kaydet
        try:
            Tk().withdraw()
            kayit_yolu = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"Tekla_Rapor_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            )
            
            if not kayit_yolu:
                return sonuclar + [⚠️ UYARI: Excel kaydedilmedi!"]
            
            # Excel yolunu sakla
            self.uretilen_excel_yolu = kayit_yolu
            
            with pd.ExcelWriter(kayit_yolu, engine='openpyxl') as writer:
                self.veri_df.to_excel(writer, sheet_name='Data', index=False)
                self.ozet_df.to_excel(writer, sheet_name='Ozet', index=False)
                
                if OPENPYXL_KURULU:
                    workbook = writer.book
                    
                    try:
                        kaynak_olusturuldu = self.kaynak_sayfa_olustur(workbook)
                        if kaynak_olusturuldu:
                            sonuclar.append(f"📋 Kaynak sayfa eklendi: '{self.dosya_adi}'")
                        else:
                            sonuclar.append("⚠️ Kaynak sayfa oluşturulamadı")
                    except Exception as e:
                        sonuclar.append(f"⚠️ Kaynak sayfa hatası: {str(e)[:50]}...")
                    
                    try:
                        data_sheet = workbook['Data']
                        
                        for col in range(1, 9):
                            data_sheet.cell(row=1, column=col).font = Font(bold=True)
                        
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        for row in data_sheet.iter_rows(min_row=1, max_row=len(self.veri_df)+1, min_col=1, max_col=8):
                            for cell in row:
                                cell.border = thin_border
                        
                        total_row = len(self.veri_df) + 3
                        data_sheet.cell(row=total_row, column=1, value="TOPLAM")
                        data_sheet.cell(row=total_row, column=7, value=self.veri_df['Total Kg'].sum())
                        data_sheet.cell(row=total_row, column=8, value=self.veri_df['Total mm'].sum())
                        
                        for col in [1, 7, 8]:
                            cell = data_sheet.cell(row=total_row, column=col)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="FFFFE6CC", end_color="FFFFE6CC", fill_type="solid")
                            cell.border = thin_border
                    except:
                        pass
                    
                    try:
                        ozet_sheet = workbook['Ozet']
                        
                        for col in range(1, 4):
                            ozet_sheet.cell(row=1, column=col).font = Font(bold=True)
                        
                        for row in ozet_sheet.iter_rows(min_row=1, max_row=len(self.ozet_df)+1, min_col=1, max_col=3):
                            for cell in row:
                                cell.border = thin_border
                        
                        ozet_total_row = len(self.ozet_df) + 3
                        ozet_sheet.cell(row=ozet_total_row, column=1, value="GENEL TOPLAM")
                        ozet_sheet.cell(row=ozet_total_row, column=2, value=self.ozet_df['Toplam mm'].sum())
                        ozet_sheet.cell(row=ozet_total_row, column=3, value=self.ozet_df['Toplam Kg'].sum())
                        
                        for col in [1, 2, 3]:
                            cell = ozet_sheet.cell(row=ozet_total_row, column=col)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="FFFFE6CC", end_color="FFFFE6CC", fill_type="solid")
                            cell.border = thin_border
                    except:
                        pass
            
            kayit_adi = os.path.basename(kayit_yolu)
            sonuclar.append(f"💾 Excel kaydedildi: {kayit_adi}")
            
            # Kesim verilerini otomatik bul
            if self.kesim_verilerini_otomatik_bul(kayit_yolu):
                sonuclar.append(f"✅ Kesim verileri otomatik bulundu: {len(self.kesim_verileri)} kayıt")
            else:
                sonuclar.append("⚠️ Kesim verileri otomatik bulunamadı")
            
            genel_toplam_kg = self.veri_df['Total Kg'].sum()
            genel_toplam_mm = self.veri_df['Total mm'].sum()
            sonuclar.append("=" * 40)
            sonuclar.append(f"📈 GENEL TOPLAM Kg: {genel_toplam_kg:,.2f}")
            sonuclar.append(f"📏 GENEL TOPLAM mm: {genel_toplam_mm:,.0f}")
            
            plaka_toplam_kg = self.veri_df[self.veri_df['Size'].str.contains('mm Sac', na=False)]['Total Kg'].sum()
            if plaka_toplam_kg > 0:
                sonuclar.append(f"🥩 PLAKA TOPLAM Kg: {plaka_toplam_kg:,.2f}")
            
            return sonuclar
                
        except Exception as e:
            return sonuclar + [f"❌ HATA (kaydetme): {str(e)}"]

# Geliştirilmiş GUI
class TeklaRaporGUI:
    def __init__(self):
        self.isleyici = TeklaRaporIsleyici()
        
        self.pencere = tk.Tk()
        self.pencere.title("Tekla Rapor İşleyici v4.0 - Kesim Optimizasyonlu")
        self.pencere.geometry("800x700")
        self.pencere.configure(bg='#f0f0f0')
        
        self.pencere.eval('tk::PlaceWindow . center')
        
        # Değişkenler
        self.stok_boy_var = StringVar(value="12000")
        self.min_fire_var = StringVar(value="400")
        
        self.arayuz_olustur()
    
    def arayuz_olustur(self):
        # Başlık
        baslik_frame = tk.Frame(self.pencere, bg='#2c3e50')
        baslik_frame.pack(fill='x', pady=(0, 10))
        
        tk.Label(baslik_frame, text="🔨 TEKLA RAPOR İŞLEYİCİ v4.0", 
                font=("Arial", 18, "bold"),
                bg='#2c3e50', fg='white').pack(pady=10)
        
        tk.Label(baslik_frame, text="PL dönüşümlü • Kesim Optimizasyonlu • Kaynak sayfalı • Otomatik İşlem", 
                font=("Arial", 9),
                bg='#2c3e50', fg='#bdc3c7').pack(pady=(0, 10))
        
        # Ana butonlar
        ana_buton_frame = tk.Frame(self.pencere, bg='#f0f0f0')
        ana_buton_frame.pack(pady=10)
        
        # Satır 1 butonları
        satir1_frame = tk.Frame(ana_buton_frame, bg='#f0f0f0')
        satir1_frame.pack()
        
        self.btn_tum_islemler = tk.Button(satir1_frame, 
                                         text="🚀 TEKLA RAPOR İŞLE",
                                         command=self.tum_islemleri_yap,
                                         bg='#e74c3c', fg='white',
                                         font=("Arial", 10, "bold"),
                                         height=2, width=20,
                                         cursor='hand2')
        self.btn_tum_islemler.pack(side='left', padx=5)
        
        # Satır 2 butonları
        satir2_frame = tk.Frame(ana_buton_frame, bg='#f0f0f0')
        satir2_frame.pack(pady=10)
        
        self.btn_benzersiz_yap = tk.Button(satir2_frame,
                                          text="🔄 BENZERSİZ YAP & KAYDET",
                                          command=self.benzersiz_yap,
                                          bg='#9b59b6', fg='white',
                                          font=("Arial", 10, "bold"),
                                          height=2, width=20,
                                          cursor='hand2')
        self.btn_benzersiz_yap.pack(side='left', padx=5)
        
        self.btn_optimizasyon = tk.Button(satir2_frame,
                                         text="⚡ OPTİMİZASYON HESAPLA",
                                         command=self.optimizasyon_hesapla,
                                         bg='#2ecc71', fg='white',
                                         font=("Arial", 10, "bold"),
                                         height=2, width=20,
                                         cursor='hand2')
        self.btn_optimizasyon.pack(side='left', padx=5)
        
        # Satır 3 butonları
        satir3_frame = tk.Frame(ana_buton_frame, bg='#f0f0f0')
        satir3_frame.pack(pady=5)
        
        self.btn_rapor_goruntule = tk.Button(satir3_frame,
                                            text="📊 RAPOR GÖRÜNTÜLE",
                                            command=self.rapor_goruntule,
                                            bg='#f39c12', fg='white',
                                            font=("Arial", 10, "bold"),
                                            height=2, width=20,
                                            cursor='hand2')
        self.btn_rapor_goruntule.pack(side='left', padx=5)
        
        self.btn_son_excel_ac = tk.Button(satir3_frame,
                                         text="📂 SON EXCEL'İ AÇ",
                                         command=self.son_excel_ac,
                                         bg='#1abc9c', fg='white',
                                         font=("Arial", 10, "bold"),
                                         height=2, width=20,
                                         cursor='hand2')
        self.btn_son_excel_ac.pack(side='left', padx=5)
        
        # Parametreler
        parametre_frame = tk.LabelFrame(self.pencere, text=" Kesim Parametreleri ",
                                       font=("Arial", 10, "bold"),
                                       padx=10, pady=10)
        parametre_frame.pack(pady=10, padx=20, fill='x')
        
        # Stok boyu
        stok_frame = tk.Frame(parametre_frame)
        stok_frame.pack(fill='x', pady=5)
        
        tk.Label(stok_frame, text="Stok Boyu (mm):", width=15, anchor='w').pack(side='left')
        tk.Entry(stok_frame, textvariable=self.stok_boy_var, width=15).pack(side='left', padx=5)
        tk.Label(stok_frame, text="mm").pack(side='left', padx=5)
        
        # Minimum fire
        fire_frame = tk.Frame(parametre_frame)
        fire_frame.pack(fill='x', pady=5)
        
        tk.Label(fire_frame, text="Minimum Fire (mm):", width=15, anchor='w').pack(side='left')
        tk.Entry(fire_frame, textvariable=self.min_fire_var, width=15).pack(side='left', padx=5)
        tk.Label(fire_frame, text="mm").pack(side='left', padx=5)
        
        # Bilgi etiketi
        bilgi_label = tk.Label(parametre_frame, 
                              text="Not: Stok boyu genellikle 12000 mm (12m), minimum fire kesilebilir artık uzunluktur.",
                              font=("Arial", 8), fg='#7f8c8d')
        bilgi_label.pack(pady=5)
        
        # Sonuç alanı
        sonuc_frame = tk.LabelFrame(self.pencere, text=" İşlem Sonuçları ",
                                   font=("Arial", 10, "bold"),
                                   padx=10, pady=10)
        sonuc_frame.pack(pady=10, padx=20, fill='both', expand=True)
        
        scrollbar = tk.Scrollbar(sonuc_frame)
        scrollbar.pack(side='right', fill='y')
        
        self.txt_sonuc = scrolledtext.ScrolledText(sonuc_frame, 
                                                  height=15,
                                                  font=("Consolas", 9),
                                                  yscrollcommand=scrollbar.set,
                                                  bg='#f8f9fa',
                                                  wrap='word')
        self.txt_sonuc.pack(side='left', fill='both', expand=True)
        
        scrollbar.config(command=self.txt_sonuc.yview)
        
        # Durum çubuğu
        self.lbl_durum = tk.Label(self.pencere, text="Hazır",
                                 bg='#2c3e50', fg='white',
                                 anchor='w', padx=10)
        self.lbl_durum.pack(side='bottom', fill='x')
    
    def log_ekle(self, mesaj):
        self.txt_sonuc.insert('end', f"{mesaj}\n")
        self.txt_sonuc.see('end')
    
    def temizle_log(self):
        self.txt_sonuc.delete('1.0', 'end')
    
    def durum_guncelle(self, mesaj):
        self.lbl_durum.config(text=mesaj)
        self.pencere.update()
    
    def butonlari_disable_et(self):
        for btn in [self.btn_tum_islemler, self.btn_benzersiz_yap,
                   self.btn_optimizasyon, self.btn_rapor_goruntule, self.btn_son_excel_ac]:
            btn.config(state='disabled')
    
    def butonlari_enable_et(self):
        for btn in [self.btn_tum_islemler, self.btn_benzersiz_yap,
                   self.btn_optimizasyon, self.btn_rapor_goruntule, self.btn_son_excel_ac]:
            btn.config(state='normal')
    
    def tum_islemleri_yap(self):
        self.temizle_log()
        self.butonlari_disable_et()
        self.durum_guncelle("Tekla rapor işleniyor...")
        
        self.log_ekle("=" * 60)
        self.log_ekle("TEKLA RAPOR İŞLEME BAŞLATILDI")
        self.log_ekle("=" * 60)
        
        sonuclar = self.isleyici.tum_islemleri_yap()
        
        for sonuc in sonuclar:
            self.log_ekle(sonuc)
        
        self.butonlari_enable_et()
        
        if any("❌ HATA" in s for s in sonuclar):
            self.durum_guncelle("Hata oluştu!")
            messagebox.showerror("Hata", "İşlem sırasında hata oluştu!")
        else:
            self.durum_guncelle("Tekla rapor işleme tamamlandı!")
            messagebox.showinfo("Başarılı", "Tekla rapor işleme tamamlandı!\nKesim verileri otomatik olarak bulundu.")
    
    def benzersiz_yap(self):
        self.temizle_log()
        self.butonlari_disable_et()
        self.durum_guncelle("Benzersizleştirme yapılıyor...")
        
        self.log_ekle("=" * 60)
        self.log_ekle("BENZERSİZLEŞTİRME İŞLEMİ")
        self.log_ekle("=" * 60)
        
        try:
            basarili = self.isleyici.benzersiz_yap()
            
            if basarili:
                self.log_ekle(f"✅ Benzersizleştirme tamamlandı: {len(self.isleyici.kesim_verileri)} benzersiz kayıt")
                self.log_ekle(f"📁 Excel'e kaydedildi: 'Benzersiz_Kesim' sayfası")
                
                # Sonuçları göster
                self.log_ekle("\nBENZERSİZ VERİLER (Büyükten küçüğe):")
                self.log_ekle("-" * 40)
                self.log_ekle("Adet\tBoy (mm)\tToplam Uzunluk")
                self.log_ekle("-" * 40)
                
                toplam_adet = 0
                toplam_uzunluk = 0
                
                for idx, row in self.isleyici.kesim_verileri.iterrows():
                    adet = int(row['Adet'])
                    boy = float(row['Boy'])
                    toplam = float(row['Toplam Uzunluk'])
                    self.log_ekle(f"{adet}\t{boy}\t{toplam:,.0f}")
                    
                    toplam_adet += adet
                    toplam_uzunluk += toplam
                
                self.log_ekle("-" * 40)
                self.log_ekle(f"TOPLAM:\t{toplam_adet}\t\t{toplam_uzunluk:,.0f}")
                
                self.durum_guncelle("Benzersizleştirme tamamlandı!")
                messagebox.showinfo("Başarılı", f"Benzersizleştirme tamamlandı!\n{len(self.isleyici.kesim_verileri)} benzersiz kayıt oluşturuldu.\nExcel'e 'Benzersiz_Kesim' sayfası olarak kaydedildi.")
            else:
                self.durum_guncelle("Benzersizleştirme başarısız!")
        
        except Exception as e:
            self.log_ekle(f"❌ HATA: {str(e)}")
            self.durum_guncelle("Hata oluştu!")
            messagebox.showerror("Hata", f"Benzersizleştirme sırasında hata: {str(e)}")
        
        finally:
            self.butonlari_enable_et()
    
    def optimizasyon_hesapla(self):
        self.temizle_log()
        self.butonlari_disable_et()
        self.durum_guncelle("Optimizasyon hesaplanıyor...")
        
        self.log_ekle("=" * 60)
        self.log_ekle("KESİM OPTİMİZASYON HESAPLAMA")
        self.log_ekle("=" * 60)
        
        try:
            # Parametreleri al
            stok_boy = float(self.stok_boy_var.get())
            min_fire = float(self.min_fire_var.get())
            
            self.log_ekle(f"📏 Stok Boyu: {stok_boy} mm")
            self.log_ekle(f"📏 Minimum Fire: {min_fire} mm")
            self.log_ekle("")
            
            sonuc = self.isleyici.kesim_optimizasyonu_hesapla(stok_boy, min_fire)
            
            if sonuc:
                self.log_ekle("✅ Optimizasyon hesaplaması tamamlandı!")
                self.log_ekle(f"📊 Kullanılan Stok Sayısı: {sonuc['kullanilan_stok_sayisi']} adet")
                self.log_ekle(f"📏 Toplam Kullanılan Uzunluk: {sonuc['toplam_kullanilan']:,.0f} mm")
                self.log_ekle(f"📏 Toplam Fire: {sonuc['toplam_fire']:,.0f} mm")
                self.log_ekle(f"📈 Ortalama Kullanım Oranı: {sonuc['ortalama_kullanim']:.1f}%")
                self.log_ekle(f"📁 Excel'e kaydedildi: 'Kesim_Optimizasyon' sayfası")
                
                # Kullanılmayan parçalar
                if sonuc['kullanilmayan_parcalar']:
                    self.log_ekle(f"⚠️ Kullanılmayan Parçalar: {len(sonuc['kullanilmayan_parcalar'])} adet")
                
                self.durum_guncelle("Optimizasyon tamamlandı!")
                messagebox.showinfo("Başarılı", f"Optimizasyon tamamlandı!\n{sonuc['kullanilan_stok_sayisi']} stok kullanıldı.\nOrtalama kullanım: {sonuc['ortalama_kullanim']:.1f}%\nExcel'e 'Kesim_Optimizasyon' sayfası olarak kaydedildi.")
            else:
                self.durum_guncelle("Optimizasyon başarısız!")
        
        except ValueError:
            self.log_ekle("❌ HATA: Geçersiz parametre değeri!")
            self.durum_guncelle("Hata oluştu!")
            messagebox.showerror("Hata", "Lütfen geçerli sayısal değerler girin!")
        except Exception as e:
            self.log_ekle(f"❌ HATA: {str(e)}")
            self.durum_guncelle("Hata oluştu!")
            messagebox.showerror("Hata", f"Optimizasyon sırasında hata: {str(e)}")
        
        finally:
            self.butonlari_enable_et()
    
    def rapor_goruntule(self):
        if self.isleyici.kesim_optimizasyon_sonucu is None:
            messagebox.showwarning("Uyarı", "Önce optimizasyon hesaplanmalı!")
            return
        
        self.temizle_log()
        self.durum_guncelle("Rapor görüntüleniyor...")
        
        try:
            rapor = self.isleyici.kesim_optimizasyon_raporu_olustur()
            
            if rapor:
                self.txt_sonuc.delete('1.0', 'end')
                self.txt_sonuc.insert('1.0', rapor)
                self.durum_guncelle("Rapor görüntülendi!")
            else:
                self.durum_guncelle("Rapor oluşturulamadı!")
        
        except Exception as e:
            self.log_ekle(f"❌ HATA: {str(e)}")
            self.durum_guncelle("Hata oluştu!")
            messagebox.showerror("Hata", f"Rapor görüntüleme sırasında hata: {str(e)}")
    
    def son_excel_ac(self):
        if self.isleyici.uretilen_excel_yolu and os.path.exists(self.isleyici.uretilen_excel_yolu):
            try:
                os.startfile(self.isleyici.uretilen_excel_yolu)  # Windows
            except:
                try:
                    import subprocess
                    subprocess.call(['open', self.isleyici.uretilen_excel_yolu])  # Mac
                except:
                    try:
                        import subprocess
                        subprocess.call(['xdg-open', self.isleyici.uretilen_excel_yolu])  # Linux
                    except:
                        messagebox.showinfo("Bilgi", f"Excel dosyası:\n{self.isleyici.uretilen_excel_yolu}")
        else:
            messagebox.showwarning("Uyarı", "Henüz Excel dosyası oluşturulmamış!")
    
    def calistir(self):
        self.pencere.mainloop()

# ANA PROGRAM
def main():
    print("Tekla Rapor İşleyici v4.0 - Otomatik Kesim Optimizasyonlu başlatılıyor...")
    
    # Kütüphane kontrolü
    try:
        import pandas as pd
        import openpyxl
        print("✓ Tüm kütüphaneler kurulu")
    except ImportError as e:
        print(f"✗ Eksik kütüphane: {e}")
        print("Lütfen şu komutları çalıştırın:")
        print("pip install pandas openpyxl")
        input("\nÇıkmak için ENTER...")
        return
    
    # GUI'yi başlat
    app = TeklaRaporGUI()
    app.calistir()

if __name__ == "__main__":

    main()
